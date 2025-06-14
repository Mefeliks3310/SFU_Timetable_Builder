import os
import re
import pandas as pd
from tkinter import messagebox, filedialog
import requests
from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class MainLogic:
    def __init__(self):
        self.teachers = []
        self.teachers_schedule = dict()

    def get_teacher_statuses(self):
        statuses = dict()
        for teacher_name in self.teachers_schedule:
            if isinstance(self.teachers_schedule[teacher_name], tuple):
                statuses[teacher_name] = "ok"
            else:
                statuses[teacher_name] = self.teachers_schedule[teacher_name]
        return statuses

    def load_config_file(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.xlsx', '.xls', '.xlsm', '.xlsb', '.csv']:
            raise ValueError("Недопустимый формат файла")

        if ext == '.csv':
            df = pd.read_csv(file_path, header=None)
        else:
            df = pd.read_excel(file_path, engine='openpyxl', header=None, names=["фио", "url"])

        if df.shape[1] != 2:
            raise ValueError("Файл конфигурации не представляет собой таблицу из двух колонок, где\n"
                             "В первой колонке - ФИО преподавателя,\n"
                             "во второй колонке - гиперссылка на преподавателя на сайте расписания СФУ.\n\n"
                             "Пожалуйста, проверьте файл конфигурации и после попробуйте загрузить ещё раз.")
        else:
            test_teachers = df.to_dict("records")
            for teacher in test_teachers:
                if "https://edu.sfu-kras.ru/timetable" not in str(teacher["url"]):
                    raise ValueError(f"Для преподавателя {teacher['фио']} введена неверная ссылка: {teacher['url']}. Убедитесь, что ссылка ведет на сайт расписания СФУ.")
            self.teachers.clear()
            self.teachers = test_teachers
            self.teachers_schedule.clear()

    def create_combined_schedule(self, check_only=False, save_file=False, callback=None, keep_groups=False):
        if not self.teachers:
            raise ValueError("Нет данных", "Сначала загрузите конфигурационный файл.")
        if check_only or not self.teachers_schedule:
            self.teachers_schedule.clear()
            for teacher in self.teachers:
                schedule = self.get_schedule(teacher["фио"], teacher["url"])
                self.teachers_schedule[teacher["фио"]] = schedule
                if callback:
                    callback(teacher["фио"], schedule)
        if check_only:
            return

        odd_schedules = {}
        even_schedules = {}
        teachers = []
        for teacher in self.teachers:
            schedule = self.teachers_schedule.get(teacher["фио"])
            if isinstance(schedule, tuple):
                teachers.append(teacher["фио"])
                odd_schedules[teacher["фио"]] = schedule[0]
                even_schedules[teacher["фио"]] = schedule[1]

        if not teachers:
            raise ValueError("Нет успешных расписаний", "Не удалось загрузить ни одно расписание. Проверьте конфигурацию.")

        odd_df = self.create_schedule_df(odd_schedules, teachers, keep_groups=keep_groups)
        even_df = self.create_schedule_df(even_schedules, teachers, keep_groups=keep_groups)
        combined_df = self.create_combined_schedule_df(odd_schedules, even_schedules, teachers, keep_groups=keep_groups)

        if not save_file:
            return

        file_path = filedialog.asksaveasfilename(
            title="Сохранить объединенное расписание",
            defaultextension=".xlsx",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if file_path:
            error_teachers = [f"{teacher}: {status}" for teacher, status in self.get_teacher_statuses().items() if status != "ok"]
            if error_teachers:
                messagebox.showwarning("Предупреждение", "Расписание создано, но есть ошибки:\n" + "\n".join(error_teachers))
            try:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    odd_df.to_excel(writer, sheet_name="1 нед", index=False)
                    even_df.to_excel(writer, sheet_name="2 нед", index=False)
                    combined_df.to_excel(writer, sheet_name="Объединённое расписание", index=False,startrow=1)
                    workbook = writer.book
                    for sheet_name in ["1 нед", "2 нед"]:
                        worksheet = writer.sheets[sheet_name]
                        self.apply_formatting(worksheet, teachers)
                    combined_worksheet = writer.sheets["Объединённое расписание"]
                    self.apply_combined_formatting(combined_worksheet, teachers)
                messagebox.showinfo("Готово", "Объединённое расписание успешно создано!")
            except PermissionError:
                raise PermissionError(f"Ошибка сохранения таблицы, возможно в данный момент открыт изменяемый файл ({file_path})")
        else:
            messagebox.showwarning("Отменено", "Сохранение файла было отменено.")


    def create_combined_schedule_df(self, odd_schedules, even_schedules, teachers, keep_groups=False):
        days_order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        rows = []
        columns = ["№", "Время"] + sum([[f"{teacher} (Нечётная)", f"{teacher} (Чётная)"] for teacher in teachers], [])

        for day in days_order:
            day_lessons = set()
            has_slots = False
            for teacher in teachers:
                if teacher in odd_schedules and day in odd_schedules[teacher]:
                    for lesson in odd_schedules[teacher][day]:
                        номер, время = lesson[0], lesson[1]
                        day_lessons.add((номер, время))
                        has_slots = True
                if teacher in even_schedules and day in even_schedules[teacher]:
                    for lesson in even_schedules[teacher][day]:
                        номер, время = lesson[0], lesson[1]
                        day_lessons.add((номер, время))
                        has_slots = True

            if has_slots:
                lessons = sorted(list(day_lessons), key=lambda x: int(x[0]))
                day_added = False
                for номер, время in lessons:
                    row = [номер, время]
                    has_data = False
                    for teacher in teachers:
                        odd_lesson = ""
                        even_lesson = ""
                        if teacher in odd_schedules and day in odd_schedules[teacher]:
                            lessons_for_time = [l for l in odd_schedules[teacher][day] if
                                                l[0] == номер and l[1] == время]
                            if lessons_for_time and lessons_for_time[0][2].strip():
                                has_data = True
                                if not keep_groups:
                                    shift_lesson_all_info = lessons_for_time[0][2].split("\n")
                                    shift_groups = shift_lesson_all_info[0]
                                    shift_groups_deleted_repeats = re.sub(
                                        r'\s*\(?(\d+\s*подгрупп[а-я]*|подгрупп[а-я]*\s*\d+)\)?\s*',
                                        '',
                                        shift_groups,
                                        flags=re.IGNORECASE
                                    ).split(", ")
                                    unique_groups_names = ", ".join(set(shift_groups_deleted_repeats))
                                    shift_lesson_all_info = unique_groups_names + "\n" + "\n".join(
                                        shift_lesson_all_info[1:])
                                    odd_lesson = shift_lesson_all_info
                                else:
                                    odd_lesson = lessons_for_time[0][2]
                        if teacher in even_schedules and day in even_schedules[teacher]:
                            lessons_for_time = [l for l in even_schedules[teacher][day] if
                                                l[0] == номер and l[1] == время]
                            if lessons_for_time and lessons_for_time[0][2].strip():
                                has_data = True
                                if not keep_groups:
                                    shift_lesson_all_info = lessons_for_time[0][2].split("\n")
                                    shift_groups = shift_lesson_all_info[0]
                                    shift_groups_deleted_repeats = re.sub(
                                        r'\s*\(?(\d+\s*подгрупп[а-я]*|подгрупп[а-я]*\s*\d+)\)?\s*',
                                        '',
                                        shift_groups,
                                        flags=re.IGNORECASE
                                    ).split(", ")
                                    unique_groups_names = ", ".join(set(shift_groups_deleted_repeats))
                                    shift_lesson_all_info = unique_groups_names + "\n" + "\n".join(
                                        shift_lesson_all_info[1:])
                                    even_lesson = shift_lesson_all_info
                                else:
                                    even_lesson = lessons_for_time[0][2]
                        row.append(odd_lesson)
                        row.append(even_lesson)
                    if has_data:
                        if not day_added:
                            rows.append([day] + [""] * (len(teachers) * 2 + 1))
                            day_added = True
                        rows.append(row)

        df = pd.DataFrame(rows, columns=columns)
        return df

    def apply_combined_formatting(self, worksheet, teachers):
        days_order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        header_fill = PatternFill(start_color="5f8a96", end_color="5f8a96", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        day_fill = PatternFill(start_color="ff6600", end_color="ff6600", fill_type="solid")
        day_font = Font(bold=True, color="FFFFFF")
        subject_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
        discipline_font = Font(color="5f8a96", bold=True)
        default_font = Font(color="000000")
        hyperlink_font = Font(color="5f8a96", bold=True, underline="single")
        medium_border = Border(left=Side(style='medium', color='cccccc'),
                               right=Side(style='medium', color='cccccc'),
                               top=Side(style='medium', color='cccccc'),
                               bottom=Side(style='medium', color='cccccc'))
        thick_bottom_border = Border(bottom=Side(style='thick', color='5f8a96'))
        day_alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)

        # Установка ширины столбцов
        worksheet.column_dimensions['A'].width = 74 / 7.5  # №
        worksheet.column_dimensions['B'].width = 100 / 7.5  # Время
        for col in range(3, worksheet.max_column + 1):
            max_length = 0
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    text = str(cell.value)
                    lines = text.split('\n')
                    for line in lines:
                        max_length = max(max_length, len(line))
            adjusted_width = max_length + 2 if max_length > 0 else 10
            worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width

        worksheet.cell(row=1, column=1).value = ""
        worksheet.cell(row=1, column=2).value = ""

        # Форматирование заголовков (строка 1 - имена преподавателей)
        for teacher_idx, teacher in enumerate(teachers):
            col_start = 3 + teacher_idx * 2  # Начало пары столбцов (Нечётная)
            col_end = col_start + 1  # Конец пары столбцов (Чётная)
            worksheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
            cell = worksheet.cell(row=1, column=col_start)
            cell.value = teacher
            cell.fill = subject_fill
            cell.font = hyperlink_font
            cell.alignment = center_alignment
            cell.border = medium_border

            # Добавляем гиперссылку если есть URL
            for t in self.teachers:
                if t["фио"] == teacher and "url" in t:
                    cell.hyperlink = t["url"]
                    # Явно переустанавливаем форматирование после гиперссылки
                    cell.font = hyperlink_font
                    cell.alignment = center_alignment
                    break

        # Форматирование строки 2 (заголовки столбцов)
        worksheet.row_dimensions[2].height = 15.75
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=2, column=col)
            if col == 1:
                cell.value = "№"
            elif col == 2:
                cell.value = "Время"
            elif col % 2 == 1:
                cell.value = "Нечётная неделя"
            else:
                cell.value = "Чётная неделя"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = medium_border
        # Форматирование данных
        for row in range(3, worksheet.max_row + 1):
            first_cell = worksheet.cell(row=row, column=1)
            if first_cell.value in days_order:
                worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=worksheet.max_column)
                for col in range(1, worksheet.max_column + 1):
                    day_cell = worksheet.cell(row=row, column=col)
                    day_cell.fill = day_fill
                    day_cell.font = day_font
                    day_cell.alignment = day_alignment
                    day_cell.border = thick_bottom_border
                    if col == 1:
                        day_cell.value = first_cell.value
            else:
                num_cell = worksheet.cell(row=row, column=1)
                time_cell = worksheet.cell(row=row, column=2)
                num_cell.alignment = center_alignment
                time_cell.alignment = center_alignment
                num_cell.border = medium_border
                time_cell.border = medium_border
                num_cell.font = default_font
                time_cell.font = default_font
                for col in range(3, worksheet.max_column + 1):
                    subject_cell = worksheet.cell(row=row, column=col)
                    if subject_cell.value:
                        lines = str(subject_cell.value).split('\n')
                        if len(lines) >= 3:
                            subject_cell.value = "\n".join(lines)
                            subject_cell.font = discipline_font
                        elif len(lines) == 2:
                            subject_cell.value = f"{lines[0]}\n{lines[1]}"
                            subject_cell.font = discipline_font
                        else:
                            subject_cell.value = lines[0]
                            subject_cell.font = default_font
                    subject_cell.fill = subject_fill
                    subject_cell.alignment = cell_alignment
                    subject_cell.border = medium_border

    def apply_formatting(self, worksheet, teachers):
        days_order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        header_fill = PatternFill(start_color="5f8a96", end_color="5f8a96", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        day_fill = PatternFill(start_color="ff6600", end_color="ff6600", fill_type="solid")
        day_font = Font(bold=True, color="FFFFFF")
        subject_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
        discipline_font = Font(color="5f8a96", bold=True)
        default_font = Font(color="000000")
        hyperlink_font = Font(color="5f8a96", bold=True)
        medium_border = Border(left=Side(style='medium', color='cccccc'),
                               right=Side(style='medium', color='cccccc'),
                               top=Side(style='medium', color='cccccc'),
                               bottom=Side(style='medium', color='cccccc'))
        thick_bottom_border = Border(bottom=Side(style='thick', color='5f8a96'))
        day_alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center", indent=1, wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)

        worksheet.column_dimensions['A'].width = 74 / 7.5
        worksheet.column_dimensions['B'].width = 100 / 7.5
        for col in range(3, worksheet.max_column + 1):
            max_length = 0
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    text = str(cell.value)
                    lines = text.split('\n')
                    for line in lines:
                        max_length = max(max_length, len(line))
            adjusted_width = max_length + 2 if max_length > 0 else 10
            worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            if col <= 2:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = medium_border
            else:
                cell.fill = subject_fill
                cell.font = hyperlink_font
                cell.alignment = cell_alignment
                cell.border = medium_border
                teacher_index = col - 3
                if teacher_index < len(teachers):
                    teacher = teachers[teacher_index]
                    for t in self.teachers:
                        if t["фио"] == teacher and "url" in t:
                            cell.hyperlink = t["url"]
                            cell.style = "Hyperlink"
                            break

        for row in range(2, worksheet.max_row + 1):
            first_cell = worksheet.cell(row=row, column=1)
            if first_cell.value in days_order:
                worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=worksheet.max_column)
                for col in range(1, worksheet.max_column + 1):
                    day_cell = worksheet.cell(row=row, column=col)
                    day_cell.fill = day_fill
                    day_cell.font = day_font
                    day_cell.alignment = day_alignment
                    day_cell.border = thick_bottom_border
                    if col == 1:
                        day_cell.value = first_cell.value
            else:
                num_cell = worksheet.cell(row=row, column=1)
                time_cell = worksheet.cell(row=row, column=2)
                num_cell.alignment = center_alignment
                time_cell.alignment = center_alignment
                num_cell.border = medium_border
                time_cell.border = medium_border
                num_cell.font = default_font
                time_cell.font = default_font
                for col in range(3, worksheet.max_column + 1):
                    subject_cell = worksheet.cell(row=row, column=col)
                    if subject_cell.value:
                        lines = str(subject_cell.value).split('\n')
                        if len(lines) >= 3:
                            subject_cell.value = "\n".join(lines)
                            subject_cell.font = discipline_font
                        elif len(lines) == 2:
                            subject_cell.value = f"{lines[0]}\n{lines[1]}"
                            subject_cell.font = discipline_font
                        else:
                            subject_cell.value = lines[0]
                            subject_cell.font = default_font
                    subject_cell.fill = subject_fill
                    subject_cell.alignment = cell_alignment
                    subject_cell.border = medium_border

    def create_schedule_df(self, schedules, teachers, keep_groups=False):
        days_order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        rows = []

        for day in days_order:
            day_lessons = set()
            for teacher in teachers:
                if teacher in schedules and day in schedules[teacher]:
                    for lesson in schedules[teacher][day]:
                        номер, время = lesson[0], lesson[1]
                        day_lessons.add((номер, время))

            if day_lessons:
                fl = True
                lessons = sorted(list(day_lessons), key=lambda x: int(x[0]))
                for номер, время in lessons:
                    row = [номер, время]
                    for teacher in teachers:
                        if teacher in schedules and day in schedules[teacher]:
                            lessons_for_time = [l for l in schedules[teacher][day] if l[0] == номер and l[1] == время]
                            if lessons_for_time and lessons_for_time[0][2].strip():
                                if not keep_groups:
                                    shift_lesson_all_info = lessons_for_time[0][2].split("\n")
                                    shift_groups = shift_lesson_all_info[0]
                                    shift_groups_deleted_repeats = re.sub(
                                        r'\s*\(?(\d+\s*подгрупп[а-я]*|подгрупп[а-я]*\s*\d+)\)?\s*',
                                        '',
                                        shift_groups,
                                        flags=re.IGNORECASE
                                    ).split(", ")
                                    unique_groups_names = ", ".join(set(shift_groups_deleted_repeats))
                                    shift_lesson_all_info = unique_groups_names + "\n" + "\n".join(shift_lesson_all_info[1:])
                                    row.append(shift_lesson_all_info)
                                else:
                                    shift_lesson_all_info = lessons_for_time[0][2].split("\n")
                                    shift_groups = shift_lesson_all_info[0].split(", ")
                                    shift_changed_groups = []
                                    for num_group in range(len(shift_groups)):
                                        if num_group % 2 != 0:
                                            shift_changed_groups.append(shift_groups[num_group] + "\n")
                                        else:
                                            if num_group + 1 == len(shift_groups):
                                                shift_changed_groups.append(shift_groups[num_group] + "\n")
                                            else:
                                                shift_changed_groups.append(shift_groups[num_group] + ", ")
                                    shift_lesson_all_info = "".join(shift_changed_groups) + "\n".join(shift_lesson_all_info[1:])
                                    row.append(shift_lesson_all_info)
                            else:
                                row.append("")
                        else:
                            row.append("")
                    if any(cell != "" for cell in row[2:]):
                        if fl:
                            rows.append([day] + [""] * (len(teachers) + 1))
                            fl = False
                        rows.append(row)

        columns = ["№", "Время"] + teachers
        df = pd.DataFrame(rows, columns=columns)
        return df

    def get_schedule(self, teacher_name, url):
        response = requests.get(url, verify=False)
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        table = soup.find('table', class_='table timetable')
        if table is None:
            return "Таблица не найдена!"
        rows = table.find_all('tr')
        if len(rows) == 0:
            return "Неправильная ссылка на расписание"
        if teacher_name[-1] != "." or teacher_name not in html:
            return "Неправильное имя преподавателя"

        odd_week_schedule = {}
        even_week_schedule = {}
        current_heading = None
        def extract_text_with_commas_and_breaks(td):
            result = []
            group_names = []
            group_set = set()
            current_line = []
            coun_br = 0
            for elem in td.contents:
                if isinstance(elem, NavigableString):
                    text_ = elem.strip()
                    if text_ and text_ != ',':
                        current_line.append(text_)
                elif isinstance(elem, Tag):
                    if elem.name == "br":
                        coun_br += 1
                        if current_line:
                            line = ' '.join(current_line).strip()
                            if line:
                                result.append(line)
                            current_line = []
                    elif elem.name == "a" and "e.sfu-kras.ru" in elem.get("href", ""):
                        result.append(elem.get("href"))
                    else:
                        text_ = elem.get_text(strip=True)
                        if "подгруппа" in text_:
                            base_name = text_
                            if base_name not in group_set:
                                group_set.add(base_name)
                                group_names.append(base_name)
                        elif coun_br == 0:
                            group_set.add(text_)
                            group_names.append(text_)
                        else:
                            current_line.append(text_)

            if current_line:
                line = ' '.join(current_line).strip()
                if line:
                    result.append(line)
            if len(result) > 2 and not (result[-2].lower().strip() == "эиос"):
                result.pop(-1)
            output = ''
            if group_names:
                output += ', '.join(group_names)
            if result:
                if output:
                    output += '\n'
                output += '\n'.join(result)
            output = output.replace(teacher_name + "\n", "")
            output = output.replace("ЭИОС\n", "ЭИОС, ")
            return output.strip()

        for row in rows:
            if "heading-section" in row.get("class", []):
                current_heading = row.get_text(separator=' ', strip=True)
                odd_week_schedule[current_heading] = []
                even_week_schedule[current_heading] = []
            if "table-center" in row.get("class", []):
                tds = row.find_all("td")
                if len(tds) > 3:
                    num_of_lesson = tds[0].get_text(separator=' ', strip=True)
                    time_of_lesson = tds[1].get_text(separator=' ', strip=True)
                    text_odd = extract_text_with_commas_and_breaks(tds[2]).strip()
                    text_even = extract_text_with_commas_and_breaks(tds[3]).strip()
                    lesson_odd = [num_of_lesson, time_of_lesson, text_odd]
                    lesson_even = [num_of_lesson, time_of_lesson, text_even]
                    odd_week_schedule[current_heading].append(lesson_odd)
                    even_week_schedule[current_heading].append(lesson_even)
                elif len(tds) == 3:
                    num_of_lesson = tds[0].get_text(separator=' ', strip=True)
                    time_of_lesson = tds[1].get_text(separator=' ', strip=True)
                    text = extract_text_with_commas_and_breaks(tds[2]).strip()
                    lesson = [num_of_lesson, time_of_lesson, text]
                    odd_week_schedule[current_heading].append(lesson)
                    even_week_schedule[current_heading].append(lesson)

        return (odd_week_schedule, even_week_schedule)